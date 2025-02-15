# 导入 Flask 及其相关模块，用于构建 Web 应用
from flask import Flask, flash, render_template, request, redirect, url_for, send_file, session
# 导入 Flask-Login 相关模块，用于处理用户登录和会话管理
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
# 再次导入 session，虽然重复但不影响功能
from flask import session
# 导入 Werkzeug 安全模块，用于密码哈希和验证
from werkzeug.security import generate_password_hash, check_password_hash
# 导入 Werkzeug 工具模块，用于安全处理文件名
from werkzeug.utils import secure_filename
# 导入 openpyxl 库，用于读取 Excel 文件
from openpyxl import load_workbook
# 导入 pymysql 游标模块，用于与 MySQL 数据库交互
import pymysql.cursors
# 从配置文件中导入配置类
from config import Config
# 再次导入 openpyxl 库，用于创建 Excel 文件
from openpyxl import Workbook
# 导入 os 模块，用于操作系统相关功能，如文件路径操作
import os
# 导入 datetime 模块，用于处理日期和时间
from datetime import datetime

# 创建 Flask 应用实例
app = Flask(__name__)
# 从配置类中加载应用配置
app.config.from_object(Config)

# 初始化 Flask-Login 的登录管理器
login_manager = LoginManager(app)
# 设置登录视图，当用户未登录访问需要登录的页面时，重定向到该视图
login_manager.login_view = 'login'

# 数据库连接
def get_db_connection():
    """
    建立与 MySQL 数据库的连接。

    :return: 返回一个 pymysql 连接对象
    """
    return pymysql.connect(
        # 从应用配置中获取 MySQL 主机地址
        host=app.config['MYSQL_HOST'],
        # 从应用配置中获取 MySQL 用户名
        user=app.config['MYSQL_USER'],
        # 从应用配置中获取 MySQL 密码
        password=app.config['MYSQL_PASSWORD'],
        # 从应用配置中获取要连接的 MySQL 数据库名
        database=app.config['MYSQL_DB'],
        # 指定游标类为 DictCursor，查询结果以字典形式返回
        cursorclass=pymysql.cursors.DictCursor
    )

class User(UserMixin):
    """
    用户类，继承自 UserMixin，用于 Flask-Login 管理用户会话。
    """
    def __init__(self, user_id):
        """
        初始化用户对象。

        :param user_id: 用户的唯一标识符
        """
        self.id = user_id

@login_manager.user_loader
def load_user(user_id):
    """
    根据用户 ID 加载用户对象，Flask-Login 会在需要时调用此函数。

    :param user_id: 用户的唯一标识符
    :return: 如果用户存在，返回 User 对象；否则返回 None
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # 从 users 表中查询指定 ID 的用户
            cursor.execute('SELECT id FROM users WHERE id = %s', (user_id,))
            user = cursor.fetchone()
            if user:
                return User(user_id)
    finally:
        # 确保无论查询结果如何，都关闭数据库连接
        conn.close()
    return None

@app.route('/logout')
@login_required  # 添加登录验证装饰器，确保只有登录用户可以访问此路由
def logout():
    """
    处理用户注销请求。

    :return: 重定向到登录页面
    """
    # 使用 Flask-Login 的 logout_user 函数注销当前用户
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """
    应用首页路由，重定向到数据导入页面。

    :return: 重定向到 import_data 路由
    """
    return redirect(url_for('import_data'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    处理用户登录请求。

    :return: 如果登录成功，重定向到首页；否则渲染登录页面
    """
    if request.method == 'POST':
        # 从表单中获取用户名和密码
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        try:
            with conn.cursor() as cursor:
                # 从 users 表中查询指定用户名的用户
                cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
                user = cursor.fetchone()
                if user and check_password_hash(user['password_hash'], password):
                    # 如果用户存在且密码验证通过，登录用户
                    login_user(User(user['id']))
                    return redirect(url_for('index'))
        finally:
            # 确保无论查询结果如何，都关闭数据库连接
            conn.close()
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """
    处理用户注销请求。

    :return: 重定向到登录页面
    """
    # 使用 Flask-Login 的 logout_user 函数注销当前用户
    logout_user()
    return redirect(url_for('login'))

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_data():
    """
    处理数据导入请求。

    :return: 如果导入成功，重定向到数据导入页面；否则渲染导入页面
    """
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
            
        file = request.files['file']
        store_id = request.form.get('store_id')
        
        if file.filename == '':
            return redirect(request.url)
            
        if file and allowed_file(file.filename):
            # 安全处理文件名
            filename = secure_filename(file.filename)
            # 构建文件保存路径
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            # 保存上传的文件
            file.save(filepath)
            
            # 读取 Excel 文件
            wb = load_workbook(filepath)
            ws = wb.active
            
            conn = get_db_connection()
            try:
                with conn.cursor() as cursor:
                    # 获取扣点比例
                    cursor.execute('SELECT deduction_rate FROM stores WHERE id = %s', (store_id,))
                    store = cursor.fetchone()
                    deduction_rate = store['deduction_rate'] if store else 0.0
                    
                    # 处理数据
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        barcode, quantity, unit_price, total = row
                        deduction_amount = total * deduction_rate
                        settlement_amount = total - deduction_amount
                        
                        # 获取商品名称
                        product_name = get_product_name(barcode)  # 需要实现商品查询逻辑
                        
                        cursor.execute('''
                            INSERT INTO sales_records 
                            (barcode, product_name, quantity, unit_price, total_amount, 
                             deduction_rate, deduction_amount, settlement_amount,
                             year_month, import_time, operator_id, store_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ''', (
                            barcode, 
                            product_name,
                            quantity,
                            unit_price,
                            total,
                            deduction_rate,
                            deduction_amount,
                            settlement_amount,
                            datetime.now().strftime('%Y-%m'),
                            datetime.now(),
                            current_user.id,
                            store_id
                        ))
                    # 提交数据库事务
                    conn.commit()
                    # 显示成功消息
                    flash('数据导入成功')
            except Exception as e:
                # 回滚数据库事务
                conn.rollback()
                # 显示错误消息
                flash(f'导入失败: {str(e)}')
            finally:
                # 确保无论操作结果如何，都关闭数据库连接
                conn.close()
            
            return redirect(url_for('import_data'))
    return render_template('import.html')

@app.route('/export')
@login_required
def export_data():
    """
    处理数据导出请求。

    :return: 如果导出成功，返回导出的 Excel 文件；否则重定向到首页
    """
    # 获取查询参数
    store_id = request.args.get('store_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # 构建基础查询
    query = '''
        SELECT s.barcode, p.name as product_name, s.quantity, 
               s.unit_price, s.total_amount, s.deduction_rate,
               s.deduction_amount, s.settlement_amount,
               s.year_month, s.import_time, 
               st.name as store_name, u.username as operator
        FROM sales_records s
        LEFT JOIN products p ON s.barcode = p.barcode
        LEFT JOIN stores st ON s.store_id = st.id
        LEFT JOIN users u ON s.operator_id = u.id
        WHERE 1=1
    '''
    params = []
    
    # 添加过滤条件
    if store_id:
        query += " AND s.store_id = %s"
        params.append(store_id)
    if start_date and end_date:
        query += " AND s.import_time BETWEEN %s AND %s" 
        params.extend([start_date, end_date])

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params)
            results = cursor.fetchall()

            # 创建Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "销售数据"
            
            # 添加表头
            headers = ['商品条码', '商品名称', '销售数量', '销售单价', '销售金额',
                      '扣点比例', '扣点金额', '结算金额', '年月', '导入时间',
                      '客户名称', '操作员']
            ws.append(headers)
            
            # 填充数据
            for row in results:
                ws.append([
                    row['barcode'],
                    row['product_name'],
                    row['quantity'],
                    row['unit_price'],
                    row['total_amount'],
                    row['deduction_rate'],
                    row['deduction_amount'],
                    row['settlement_amount'],
                    row['year_month'],
                    row['import_time'].strftime('%Y-%m-%d %H:%M:%S'),
                    row['store_name'],
                    row['operator']
                ])
            
            # 保存临时文件
            filename = f"export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
            wb.save(filepath)
            
            return send_file(filepath, as_attachment=True, download_name=f"销售数据导出_{filename}")
            
    except Exception as e:
        flash(f'导出失败: {str(e)}')
        return redirect(url_for('index'))
    finally:
        conn.close()

def get_product_name(barcode):
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute('SELECT name FROM products WHERE barcode = %s', (barcode,))
            product = cursor.fetchone()
            if product:
                return product['name']
    finally:
        conn.close()
    return None

if __name__ == '__main__':
    app.run(debug=True)
